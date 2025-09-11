import os
import re
from datetime import datetime
import openpyxl
import win32com.client
import config
import traceback

def create_order_pdf(supplier_name, items, sales_contact, sender_info):
    """
    Excelテンプレートから注文書PDFを作成する。
    この関数はwin32comを使用するため、Windows環境とExcelのインストールが必要です。
    また、スレッドから呼び出す場合は、呼び出し元でpythoncom.CoInitialize()と
    CoUninitialize()を処理する必要があります。
    """
    excel = None
    workbook = None
    temp_excel_path = None
    
    # テンプレートファイルの存在チェック
    if not config.EXCEL_TEMPLATE_PATH or not os.path.exists(config.EXCEL_TEMPLATE_PATH):
        print(f"エラー: Excelテンプレートが見つかりません。パスを確認してください: {config.EXCEL_TEMPLATE_PATH}")
        return None

    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        # テンプレートを開き、データを書き込む
        wb = openpyxl.load_workbook(config.EXCEL_TEMPLATE_PATH)
        ws = wb.active
        ws["A5"] = f"{supplier_name} 御中"
        ws["A7"] = f"{sales_contact} 様"
        ws["D8"] = f"担当：{sender_info['name']}"
        ws["D14"] = sender_info["email"]

        for i, item in enumerate(items):
            ws[f"A{16+i}"] = item["db_part_number"]
            ws[f"B{16+i}"] = item["maker_name"]
            ws[f"C{16+i}"] = item["quantity"]
        
        # PDFのファイルパスを生成
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        safe_supplier_name = re.sub(r'[\\/:*?"<>|]', '_', supplier_name)
        pdf_filename = f"{timestamp}_{safe_supplier_name}_注文書.pdf"
        
        if not os.path.exists(config.PDF_SAVE_DIR):
            os.makedirs(config.PDF_SAVE_DIR)
        pdf_path = os.path.join(config.PDF_SAVE_DIR, pdf_filename)
        
        # 一時的なExcelファイルとして保存
        temp_excel_path = os.path.join(config.PDF_SAVE_DIR, f"temp_{timestamp}.xlsx")
        wb.save(temp_excel_path)
        wb.close()

        # win32comでPDFに変換
        workbook = excel.Workbooks.Open(temp_excel_path)
        workbook.ActiveSheet.ExportAsFixedFormat(0, pdf_path)
        
        print(f"-> PDF作成完了: {pdf_filename}")
        return pdf_path
        
    except Exception as e:
        print(f"PDF作成エラー ({supplier_name}): {e}")
        traceback.print_exc()
        return None
        
    finally:
        # クリーンアップ処理
        if workbook:
            workbook.Close(SaveChanges=False)
        if excel:
            excel.Quit()
        if temp_excel_path and os.path.exists(temp_excel_path):
            try:
                os.remove(temp_excel_path)
            except Exception as e:
                print(f"一時ファイルの削除に失敗: {e}")
