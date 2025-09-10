import os
import re
import sys
import time
import queue
import threading
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from collections import defaultdict
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import scrolledtext, messagebox, ttk

import openpyxl
import win32com.client # PDF作成時のExcel操作にのみ使用
from dotenv import load_dotenv, dotenv_values
from notion_client import Client

# .envファイルから環境変数を読み込む
load_dotenv()

# --- 定数 ---
NOTION_API_TOKEN = os.getenv("NOTION_API_TOKEN")
PAGE_ID_CONTAINING_DB = os.getenv("NOTION_DATABASE_ID")
EXCEL_TEMPLATE_PATH = r"C:\Users\SEIZOU-20\Desktop\注文書.xlsx"
PDF_SAVE_DIR = os.path.join(Path.home(), "Desktop", "注文書")

# SMTPサーバー情報 (アカウント情報はアプリ内で動的に読み込み)
SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.office365.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))


class QueueIO:
    def __init__(self, q): self.q = q
    def write(self, text): self.q.put(("log", text))
    def flush(self): pass

def get_order_data_from_notion():
    if not NOTION_API_TOKEN or not PAGE_ID_CONTAINING_DB: return None
    notion = Client(auth=NOTION_API_TOKEN)
    order_list = []
    try:
        print(f"DBコンテナ ({PAGE_ID_CONTAINING_DB}) を検索中...")
        children = notion.blocks.children.list(block_id=PAGE_ID_CONTAINING_DB)
        real_database_id = next((b.get("id") for b in children.get("results", []) if b.get("type") == "child_database"), None)
        if not real_database_id: print(f"エラー: DBが見つかりません。"); return []
        print(f"DB発見: {real_database_id}")
        
        all_results = []
        next_cursor = None
        while True:
            query_res = notion.databases.query(database_id=real_database_id, start_cursor=next_cursor)
            all_results.extend(query_res.get("results", []))
            if not query_res.get("has_more"): break
            next_cursor = query_res.get("next_cursor")

        print(f"全 {len(all_results)} 件のデータをフィルタリング中...")
        for page in all_results:
            props = page.get("properties", {})
            if "要発注" not in props.get("注文ステータス", {}).get("formula", {}).get("string", ""): continue
            
            supplier_relation = props.get("DB_仕入先マスター", {}).get("relation", [])
            if not supplier_relation: continue
            supplier_page_id = supplier_relation[0].get("id")

            try:
                time.sleep(0.35)
                supplier_page = notion.pages.retrieve(page_id=supplier_page_id)
                supplier_props = supplier_page.get("properties", {})
                order_list.append({
                    "page_id": page["id"],
                    "maker_name": props.get("メーカー名", {}).get("rich_text", [{}])[0].get("plain_text", ""),
                    "db_part_number": props.get("DB品番", {}).get("rich_text", [{}])[0].get("plain_text", ""),
                    "quantity": props.get("数量", {}).get("number", 0),
                    "supplier_name": supplier_props.get("購入先", {}).get("title", [{}])[0].get("plain_text", ""),
                    "sales_contact": supplier_props.get("営業担当", {}).get("rich_text", [{}])[0].get("plain_text", ""),
                    "email": supplier_props.get("メール", {}).get("email", ""),
                    "email_cc": supplier_props.get("メール CC:", {}).get("email", ""),
                })
            except Exception as e: print(f"仕入先情報取得エラー (Page ID: {supplier_page_id}): {e}")
        print(f"-> フィルタリング完了。{len(order_list)} 件の要発注データが見つかりました。")
    except Exception as e: print(f"Notion DB処理エラー: {e}")
    return order_list

def create_order_pdf(supplier_name, items, sales_contact, sender_info):
    excel_app = None
    try:
        excel_app = win32com.client.Dispatch("Excel.Application")
        excel_app.Visible = False
        wb = openpyxl.load_workbook(EXCEL_TEMPLATE_PATH)
        ws = wb.active
        ws["A5"] = f"{supplier_name} 御中"; ws["A7"] = f"{sales_contact} 様"
        ws["D8"] = f"担当：{sender_info['name']}"
        ws["D14"] = sender_info["email"]

        for i, item in enumerate(items): ws[f"A{16+i}"], ws[f"B{16+i}"], ws[f"C{16+i}"] = item["db_part_number"], item["maker_name"], item["quantity"]
        
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        safe_supplier_name = re.sub(r'''[\\/:*?"<>|]''', '_', supplier_name)
        pdf_filename = f"{timestamp}_{safe_supplier_name}_注文書.pdf"
        if not os.path.exists(PDF_SAVE_DIR): os.makedirs(PDF_SAVE_DIR)
        pdf_path = os.path.join(PDF_SAVE_DIR, pdf_filename)
        temp_excel_path = os.path.join(PDF_SAVE_DIR, f"temp_{timestamp}.xlsx")
        
        wb.save(temp_excel_path)
        wb.close()

        workbook = excel_app.Workbooks.Open(temp_excel_path)
        workbook.ActiveSheet.ExportAsFixedFormat(0, pdf_path)
        workbook.Close(False)
        os.remove(temp_excel_path)
        print(f"-> PDF作成完了: {pdf_filename}")
        return pdf_path
    except Exception as e: print(f"PDF作成エラー ({supplier_name}): {e}"); return None
    finally: 
        if excel_app: excel_app.Quit()

def send_smtp_mail(info, pdf_path, sender_creds):
    try:
        msg = MIMEMultipart()
        msg["From"] = sender_creds["sender"]
        msg["To"] = info["email"]
        if info["email_cc"]: msg["Cc"] = info["email_cc"]
        msg["Subject"] = "注文書送付の件"

        body = f'''{info["supplier_name"]}\n{info.get('sales_contact', 'ご担当者')} 様\n\nいつも大変お世話になります。\n添付の通り注文宜しくお願い致します。\n\n∝∝∝∝∝∝∝∝∝∝∝∝∝∝∝∝\n株式会社　新井精密\n製造課　発注担当\n〒368-0061\n埼玉県秩父市小柱670番地\nTEL: 0494-26-7786\nFAX: 0494-26-7787\n∝∝∝∝∝∝∝∝∝∝∝∝∝∝∝∝'''
        msg.attach(MIMEText(body, 'plain'))

        with open(pdf_path, "rb") as f:
            part = MIMEApplication(f.read(), Name=os.path.basename(pdf_path))
        part['Content-Disposition'] = f'attachment; filename="{os.path.basename(pdf_path)}"'
        msg.attach(part)

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(sender_creds["sender"], sender_creds["password"])
            recipients = [info["email"]] + ([info["email_cc"]] if info["email_cc"] else [])
            server.sendmail(sender_creds["sender"], recipients, msg.as_string())
        print(f"-> メール送信完了 (From: {sender_creds['sender']})")
        return True
    except Exception as e: print(f"メール送信エラー: {e}"); return False

def update_notion_pages(page_ids):
    print(f"{len(page_ids)}件の「発注日」を更新中...")
    notion = Client(auth=NOTION_API_TOKEN)
    today = datetime.now().strftime("%Y-%m-%d")
    for i, page_id in enumerate(page_ids):
        try:
            notion.pages.update(page_id=page_id, properties={"発注日": {"date": {"start": today}}})
            print(f"({i+1}/{len(page_ids)}) {page_id} 更新完了")
            time.sleep(0.35)
        except Exception as e: print(f"エラー: {page_id} 更新失敗. {e}")
    print("Notion更新完了")

class Application(ttk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.master.title("Notion注文書メール作成アプリ")
        self.q = queue.Queue(); self.queue_io = QueueIO(self.q)
        self.processing = False; self.order_data = []; self.current_pdf_path = None; self.sent_suppliers = set()
        self.accounts = {}; self.selected_account = tk.StringVar()
        self.load_accounts_from_env()
        self.configure_styles()
        self.create_widgets()
        self.master.protocol("WM_DELETE_WINDOW", self.on_closing)

    def load_accounts_from_env(self):
        config = dotenv_values()
        sender_prefix = "EMAIL_SENDER_"
        sender_keys = [key for key in config if key.startswith(sender_prefix)]
        for key in sender_keys:
            account_name = key[len(sender_prefix):]
            password_key = f"EMAIL_PASSWORD_{account_name}"
            sender_email = config.get(key)
            password = config.get(password_key)
            if sender_email and password:
                self.accounts[account_name] = {"sender": sender_email, "password": password}
        if self.accounts:
            account_names = list(self.accounts.keys())
            self.selected_account.set(account_names[0])

    def configure_styles(self):
        style = ttk.Style(self.master); style.theme_use("clam")
        self.BG_COLOR = "#E8F0F9"; self.TEXT_COLOR = "#00224D"; self.BUTTON_BG = "#4A90E2"; self.BUTTON_FG = "#FFFFFF"; self.PROGRESS_BG = "#4A90E2"; self.HEADER_BG = "#357ABD";         self.EMPHASIS_COLOR = "#007BFF"
        style.configure(".", background=self.BG_COLOR, foreground=self.TEXT_COLOR, font=("Yu Gothic UI", 10))
        style.configure("TFrame", background=self.BG_COLOR)
        style.configure("TLabel", background=self.BG_COLOR, foreground=self.TEXT_COLOR, font=("Yu Gothic UI", 10))
        style.configure("TLabelFrame", background=self.BG_COLOR, bordercolor=self.HEADER_BG)
        style.configure("TLabelFrame.Label", background=self.BG_COLOR, foreground=self.TEXT_COLOR, font=("Yu Gothic UI", 12, "bold"))
        style.configure("TButton", background=self.BUTTON_BG, foreground=self.BUTTON_FG, font=("Yu Gothic UI", 11, "bold"), borderwidth=0, padding=5)
        style.map("TButton", background=[("active", "#357ABD"), ("disabled", "#C0C0C0")])
        style.configure("Treeview", background="#FFFFFF", fieldbackground="#FFFFFF", foreground=self.TEXT_COLOR, rowheight=25, font=("Yu Gothic UI", 9))
        style.map("Treeview", background=[("selected", "#4A90E2")], foreground=[("selected", "#FFFFFF")])
        style.configure("Treeview.Heading", background=self.HEADER_BG, foreground=self.BUTTON_FG, font=("Yu Gothic UI", 10, "bold"))
        style.map("Treeview.Heading", background=[("active", "#4A90E2")])
        style.configure("TCombobox", font=("Yu Gothic UI", 10))

    def create_widgets(self):
        self.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # --- Top Action Pane ---
        top_pane = ttk.Frame(self, padding=10); top_pane.pack(fill=tk.X)
        account_frame = ttk.LabelFrame(top_pane, text="送信者アカウント"); account_frame.pack(side=tk.LEFT, padx=(0, 20), fill=tk.X, expand=True)
        self.account_selector = ttk.Combobox(account_frame, textvariable=self.selected_account, values=sorted(list(self.accounts.keys())), state="readonly", width=30, font=("Yu Gothic UI", 11)); self.account_selector.pack(padx=10, pady=10, fill=tk.X)
        self.get_data_button = ttk.Button(top_pane, text="1. Notionからデータを取得", command=self.start_data_retrieval); self.get_data_button.pack(side=tk.LEFT, ipady=5, ipadx=5, padx=10)

        # --- Middle Content Pane ---
        middle_pane = ttk.PanedWindow(self, orient=tk.HORIZONTAL); middle_pane.pack(fill=tk.BOTH, expand=True, pady=10)
        
        supplier_pane = ttk.Frame(middle_pane, padding=5); middle_pane.add(supplier_pane, weight=1)
        table_pane = ttk.Frame(middle_pane, padding=5); middle_pane.add(table_pane, weight=2)

        supplier_frame = ttk.LabelFrame(supplier_pane, text="2. 仕入先リスト"); supplier_frame.pack(fill=tk.BOTH, expand=True)
        self.supplier_listbox = tk.Listbox(supplier_frame, exportselection=False, bg="#FFFFFF", fg=self.TEXT_COLOR, selectbackground=self.HEADER_BG, selectforeground=self.BUTTON_FG, font=("Yu Gothic UI", 11))
        self.supplier_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True); self.supplier_listbox.bind("<<ListboxSelect>>", self.on_supplier_select)
        supplier_vsb = ttk.Scrollbar(supplier_frame, orient="vertical", command=self.supplier_listbox.yview); self.supplier_listbox.config(yscrollcommand=supplier_vsb.set); supplier_vsb.pack(side=tk.RIGHT, fill=tk.Y)

        table_frame = ttk.LabelFrame(table_pane, text="発注対象データ (選択した仕入先)"); table_frame.pack(fill=tk.BOTH, expand=True)
        columns = ("maker", "part_num", "qty"); self.tree = ttk.Treeview(table_frame, columns=columns, show="headings")
        self.tree.heading("maker", text="メーカー"); self.tree.heading("part_num", text="品番"); self.tree.heading("qty", text="数量")
        self.tree.column("maker", width=150); self.tree.column("part_num", width=250); self.tree.column("qty", width=60, anchor=tk.E)
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview); hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set); vsb.pack(side=tk.RIGHT, fill=tk.Y); hsb.pack(side=tk.BOTTOM, fill=tk.X); self.tree.pack(fill=tk.BOTH, expand=True)

        # --- Bottom Action Pane ---
        bottom_pane = ttk.Frame(self, padding=10); bottom_pane.pack(fill=tk.X)
        preview_frame = ttk.LabelFrame(bottom_pane, text="3. 内容を確認して送信"); preview_frame.pack(fill=tk.X, expand=True, side=tk.LEFT, padx=(0, 20))
        self.to_var, self.cc_var, self.contact_var, self.pdf_var = tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar()
        ttk.Label(preview_frame, text="宛先(To):").grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
        ttk.Label(preview_frame, textvariable=self.to_var).grid(row=0, column=1, sticky=tk.W, padx=5)
        ttk.Label(preview_frame, text="宛先(Cc):").grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
        ttk.Label(preview_frame, textvariable=self.cc_var).grid(row=1, column=1, sticky=tk.W, padx=5)
        ttk.Label(preview_frame, text="担当者:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=2)
        ttk.Label(preview_frame, textvariable=self.contact_var).grid(row=2, column=1, sticky=tk.W, padx=5)
        ttk.Label(preview_frame, text="添付PDF:").grid(row=3, column=0, sticky=tk.W, padx=5, pady=2)
        self.pdf_label = ttk.Label(preview_frame, textvariable=self.pdf_var, foreground="blue", cursor="hand2"); self.pdf_label.grid(row=3, column=1, sticky=tk.W, padx=5)
        self.pdf_label.bind("<Button-1>", lambda e: self.open_current_pdf())

        send_button_frame = ttk.Frame(bottom_pane); send_button_frame.pack(side=tk.LEFT, expand=True, fill=tk.BOTH)
        self.send_mail_button = ttk.Button(send_button_frame, text="メール送信", command=self.send_single_mail, state="disabled"); self.send_mail_button.pack(expand=True, ipady=10, ipadx=10)

        # --- Log Pane ---
        log_frame = ttk.LabelFrame(self, text="ログ"); log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(5,0))
        self.log_display = scrolledtext.ScrolledText(log_frame, height=5, wrap=tk.WORD, bg="#FFFFFF", fg=self.TEXT_COLOR, font=("Yu Gothic UI", 12)); self.log_display.pack(fill=tk.BOTH, expand=True); self.log_display.configure(state='disabled')
        self.log_display.tag_configure("emphasis", foreground=self.EMPHASIS_COLOR, font=("Yu Gothic UI", 12, "bold"))

    def start_data_retrieval(self):
        if self.processing: return
        if not all([NOTION_API_TOKEN, self.accounts]): return messagebox.showerror("設定エラー", ".envファイルにNotionとEmailアカウント(EMAIL_SENDER_xx)の設定が必要です。")
        self.processing = True; self.toggle_buttons(False); self.clear_displays()
        threading.Thread(target=self.run_thread, args=(self.get_data_task,)).start()
        self.master.after(100, self.check_queue)

    def on_supplier_select(self, event=None):
        if self.processing or not self.supplier_listbox.curselection(): return
        selected_supplier = self.supplier_listbox.get(self.supplier_listbox.curselection())
        if selected_supplier in self.sent_suppliers: self.clear_preview(); return
        self.processing = True; self.toggle_buttons(False)
        self.update_table_for_supplier(selected_supplier)
        threading.Thread(target=self.run_thread, args=(self.create_pdf_task,)).start()
        self.master.after(100, self.check_queue)

    def send_single_mail(self):
        if self.processing or not self.current_pdf_path: return
        if not messagebox.askyesno("メール送信確認", f"{self.to_var.get()} 宛にメールを送信します。よろしいですか？"): return
        self.processing = True; self.toggle_buttons(False)
        threading.Thread(target=self.run_thread, args=(self.send_mail_task,)).start()
        self.master.after(100, self.check_queue)

    def run_thread(self, task_func, *args):
        original_stdout = sys.stdout; sys.stdout = self.queue_io
        try: task_func(*args)
        except Exception as e: print(f"\n予期せぬエラー: {e}")
        finally: sys.stdout = original_stdout; self.q.put(("task_complete", None))

    def get_data_task(self): print("Notionからデータ取得中..."); data = get_order_data_from_notion(); self.q.put(("update_data_ui", data))
    def create_pdf_task(self):
        account_name = self.selected_account.get()
        sender_creds = self.accounts[account_name]
        sender_info = {"name": account_name, "email": sender_creds["sender"]}
        selected_supplier = self.supplier_listbox.get(self.supplier_listbox.curselection())
        print(f"「{selected_supplier}」のPDFを作成中...")
        items = [item for item in self.order_data if item["supplier_name"] == selected_supplier]
        pdf_path = create_order_pdf(selected_supplier, items, items[0]["sales_contact"], sender_info)
        self.q.put(("update_preview_ui", (items[0], pdf_path)))

    def send_mail_task(self):
        account_name = self.selected_account.get()
        sender_creds = self.accounts[account_name]
        selected_supplier = self.supplier_listbox.get(self.supplier_listbox.curselection())
        print(f"「{selected_supplier}」宛にメールを送信中 (From: {sender_creds['sender']})...")
        items = [item for item in self.order_data if item["supplier_name"] == selected_supplier]
        success = send_smtp_mail(items[0], self.current_pdf_path, sender_creds)
        if success:
            page_ids_to_update = [item['page_id'] for item in items]
            self.q.put(("ask_and_update_notion", (selected_supplier, page_ids_to_update)))

    def check_queue(self):
        try:
            while True:
                command, data = self.q.get_nowait()
                if command == "log": self.log(data.strip())
                elif command == "update_data_ui": self.update_data_ui(data)
                elif command == "update_preview_ui": self.update_preview_ui(data)
                elif command == "ask_and_update_notion": self.ask_and_update_notion(data[0], data[1])
                elif command == "mark_as_sent_after_update": self.mark_as_sent(data)
                elif command == "task_complete": self.processing = False; self.toggle_buttons(True)
        except queue.Empty: self.master.after(100, self.check_queue)

    def log(self, message, tag=None):
        self.log_display.config(state="normal")
        if tag:
            self.log_display.insert(tk.END, message + "\n", tag)
        else:
            self.log_display.insert(tk.END, message + "\n")
        self.log_display.see(tk.END)
        self.log_display.config(state="disabled")

    def update_data_ui(self, data):
        self.order_data = data; self.sent_suppliers.clear()
        self.tree.delete(*self.tree.get_children())
        self.update_supplier_list(data)
        if not data: self.log("-> 発注対象のデータは見つかりませんでした。")
        else: self.log("-> データ取得完了。左のリストから仕入先を選択してください。", "emphasis")

    def update_preview_ui(self, data):
        info, pdf_path = data
        self.current_pdf_path = pdf_path
        self.to_var.set(info.get("email", "")); self.cc_var.set(info.get("email_cc", "")); self.contact_var.set(info.get("sales_contact", ""))
        self.pdf_var.set(os.path.basename(pdf_path) if pdf_path else "作成失敗")
        self.log("-> プレビューの準備ができました。")
        self.log("   宛先、担当者、PDFの内容を必ず確認してください。", "emphasis")
        self.log("   (PDFファイル名をクリックすると内容を開けます)", "emphasis")
        self.log("-> 問題がなければ「メール送信」ボタンを押してください。", "emphasis")

    def ask_and_update_notion(self, supplier, page_ids):
        if messagebox.askyesno("Notion更新確認", f"メール送信が完了しました。\n\n「{supplier}」のNotionページの「発注日」を更新しますか？"):
            self.processing = True; self.toggle_buttons(False)
            self.log(f"「{supplier}」のNotionページを更新中...")
            threading.Thread(target=self.run_thread, args=(self.update_notion_task, page_ids)).start()
        else: self.mark_as_sent(supplier, updated=False)

    def update_notion_task(self, page_ids):
        update_notion_pages(page_ids)
        supplier = next((item["supplier_name"] for item in self.order_data if item["page_id"] == page_ids[0]), None)
        if supplier: self.q.put(("mark_as_sent_after_update", supplier))

    def mark_as_sent(self, supplier, updated=True):
        self.sent_suppliers.add(supplier)
        try:
            idx = self.supplier_listbox.get(0, "end").index(supplier)
            self.supplier_listbox.itemconfig(idx, {'fg': 'gray'}); self.supplier_listbox.selection_clear(idx)
        except ValueError: pass
        self.clear_preview()
        status = "更新済み" if updated else "更新スキップ"
        self.log(f"-> 「{supplier}」は送信済みとしてマークされました。({status})")

    def update_table_for_supplier(self, supplier_name):
        self.tree.delete(*self.tree.get_children())
        items_to_display = [item for item in self.order_data if item["supplier_name"] == supplier_name]
        for item in items_to_display:
            self.tree.insert("", tk.END, values=(item.get("maker_name", ""), item.get("db_part_number", ""), item.get("quantity", 0)))

    def update_supplier_list(self, data):
        self.supplier_listbox.delete(0, tk.END)
        suppliers = sorted(list(set(i["supplier_name"] for i in data)))
        for s in suppliers: self.supplier_listbox.insert(tk.END, s)

    def clear_displays(self): 
        self.tree.delete(*self.tree.get_children()); self.supplier_listbox.delete(0, tk.END); self.sent_suppliers.clear(); self.clear_preview()
        self.log_display.config(state="normal"); self.log_display.delete(1.0, tk.END); self.log_display.config(state="disabled")

    def clear_preview(self): self.to_var.set(""); self.cc_var.set(""); self.contact_var.set(""); self.pdf_var.set(""); self.current_pdf_path = None; self.send_mail_button.config(state="disabled")
    def toggle_buttons(self, enabled): 
        state = "normal" if enabled else "disabled"
        self.get_data_button.config(state=state)
        self.send_mail_button.config(state="disabled")
        if enabled and self.current_pdf_path: self.send_mail_button.config(state="normal")

    def open_current_pdf(self): 
        if self.current_pdf_path and os.path.exists(self.current_pdf_path): os.startfile(self.current_pdf_path)
        else: messagebox.showwarning("ファイルなし", "PDFファイルが見つかりません。")

    def on_closing(self):
        if self.processing: return messagebox.showwarning("処理中", "処理が実行中です。終了できません。")
        self.master.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    root.state('zoomed')
    app = Application(master=root)
    app.mainloop()